import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

# ========= ПАРАМЕТРЫ =========

OLD_PATH = "3-semestr/Office_Tech/Lab_3/schedule/iait_22-23-2.02.03.xlsx"
NEW_PATH = "3-semestr/Office_Tech/Lab_3/schedule/iait_22-23-2.03.10.xlsx"
RESULT_PATH = "3-semestr/resources/out/3.1/разница.xlsx"

FIRST_DATA_ROW = 5

COLUMNS_ODD  = ("D", "E", "F", "G", "H")
COLUMNS_EVEN = ("K", "L", "M", "N", "O")

STYLE_REMOVED = PatternFill(fill_type="solid", fgColor="FFC7CE")
STYLE_UPDATED = PatternFill(fill_type="solid", fgColor="FFEB9C")
STYLE_NEW     = PatternFill(fill_type="solid", fgColor="C6EFCE")


# ========= СЕРВИСНЫЕ ФУНКЦИИ =========

def clean(value):
    return "" if value is None else str(value).strip()


def collect_entries(sheet):
    result = {}

    group = None
    weekday = None

    for r in range(FIRST_DATA_ROW, sheet.max_row + 1):
        if sheet[f"A{r}"].value:
            group = clean(sheet[f"A{r}"].value)

        if sheet[f"B{r}"].value:
            weekday = clean(sheet[f"B{r}"].value)

        pair = clean(sheet[f"C{r}"].value)

        if sheet[f"D{r}"].value:
            result[(group, weekday, pair, 1)] = (r, COLUMNS_ODD)

        if sheet[f"K{r}"].value:
            result[(group, weekday, pair, 2)] = (r, COLUMNS_EVEN)

    return result


def resize_row(sheet, row, columns):
    max_lines = 1

    for c in columns:
        txt = sheet[f"{c}{row}"].value
        if txt:
            lines = len(str(txt)) // 30 + 1
            max_lines = max(max_lines, lines)

    sheet.row_dimensions[row].height = max(20, max_lines * 15)


# ========= ОСНОВНОЙ ПРОЦЕСС =========

wb_old = openpyxl.load_workbook(OLD_PATH)
wb_new = openpyxl.load_workbook(NEW_PATH)
wb_result = openpyxl.load_workbook(NEW_PATH)

for sh_old, sh_new, sh_res in zip(
        wb_old.worksheets,
        wb_new.worksheets,
        wb_result.worksheets):

    old_data = collect_entries(sh_old)
    new_data = collect_entries(sh_new)

    old_keys = set(old_data.keys())
    new_keys = set(new_data.keys())

    # --- УДАЛЁННЫЕ ЗАНЯТИЯ ---
    for key in old_keys - new_keys:
        row, cols = old_data[key]
        for col in cols:
            cell = sh_res[f"{col}{row}"]
            cell.value = sh_old[f"{col}{row}"].value
            cell.fill = STYLE_REMOVED

    # --- ДОБАВЛЕННЫЕ ЗАНЯТИЯ ---
    for key in new_keys - old_keys:
        row, cols = new_data[key]
        for col in cols:
            sh_res[f"{col}{row}"].fill = STYLE_NEW

    # --- ИЗМЕНЁННЫЕ ЗАНЯТИЯ ---
    for key in old_keys & new_keys:
        old_row, _ = old_data[key]
        new_row, cols = new_data[key]

        modified = False

        for col in cols:
            before = clean(sh_old[f"{col}{old_row}"].value)
            after = clean(sh_new[f"{col}{new_row}"].value)

            if before != after:
                cell = sh_res[f"{col}{new_row}"]
                cell.value = f"{before} -> {after}"
                cell.fill = STYLE_UPDATED

                align = copy(cell.alignment)
                align.wrap_text = True
                cell.alignment = align

                modified = True

        if modified:
            resize_row(sh_res, new_row, cols)

wb_result.save(RESULT_PATH)
